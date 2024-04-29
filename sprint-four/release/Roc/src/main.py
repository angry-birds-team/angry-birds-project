import os
os.environ['OPENCV_FFMPEG_READ_ATTEMPTS'] = "2147483647"
print(os.environ['OPENCV_FFMPEG_READ_ATTEMPTS'])
import cv2
import numpy as np
import tensorflow as tf
import openpyxl
import json
import time
from tkinter import Tk, filedialog, Button, Label, Entry, Text, Scrollbar
from tkinter import *
from tkinter import ttk
import tkinter as tk
from PIL import Image, ImageTk
from pathlib import Path
from collections import deque

# enable these if deploying, otherwise, disable
path = os.path.dirname(os.path.abspath(__file__))
path = os.path.dirname(path)
path = os.path.dirname(path)
path = os.path.dirname(path)
os.chdir(path)
print(path)
print(os.getcwd())
parent = os.getcwd()
# define paths for project
# remember, normal current cwd is angry-birds-project
src_path = "src/"
models_path = "models/"
input_path = "input/"
output_path = "output/"


# load configuration settings
f = open(f'{src_path}config.json', 'r+')
settings = json.load(f)
model_selected = settings["model_selected"]
wren_model_path = settings["wren_model_path"]
warbler_model_path = settings["warbler_model_path"]
input_video_path = settings["input_video_path"] # no longer needed, i believe?  (video files are selected in open_file)
output_video_path = settings["output_video_path"]
output_timestamps_path = settings["output_timestamps_path"]
frame_divisor = int(settings["frame_divisor"]) # Only frames with a frame number divisible by this number will be processed (1 for all frames, this is for optimization) 
confidence_threshold = float(settings["confidence_threshold"]) # confidence threshold should be between 0 and 1
model_int = 0

if model_selected == "wren":
    model_int = 1
    input_model_path = wren_model_path
else:
    model_int = 2
    input_model_path = warbler_model_path

# create interpreter and load with pre-trained model 
interpreter = tf.lite.Interpreter(model_path=input_model_path)
interpreter.allocate_tensors()


input_details = interpreter.get_input_details() # list of dictionaries, each dictionary has details about an input tensor
output_details = interpreter.get_output_details() # list of dictionaries, each dictionary has details about an input tensor
input_shape = input_details[0]['shape'] # array of shape of input tensor

# Declare global variables

# variable for whether or not the video is playback is paused. True for playing, False for Paused
playing = False
# variable for current frame being read. initialize when program is run and reset when a new video is loaded
frame_number = 0
cap = None
confidence_percentage = 0
timestamp_start = None
time_position = ""
formatting_lines = ["Start Time (min:sec) - End Time (min:sec)"]

# initialize variables for delay to avoid duplicate timestamps
# don't set to above zero, this code isn't working yet
delay_started = False
delay_start_time = None
delay_duration = 0  # delay duration in seconds (how long the bird needs to be missing to end a timestamp)

# Initialize variables for video selection
input_video_paths = deque()
current_video_index = 0
video_selected = ""

# open workbook for collecting timestamps to later output to excel file
# reserve first sheet for first video
global_workbook = openpyxl.Workbook()

def open_files(): #function to a set of files
    global input_video_paths
    global cap
    global frame_number
    global sheet

    # open system dialog to open video files
    paths_tuple = filedialog.askopenfilenames(title="Select Video Files", filetypes=(("MP4 files", "*.mp4"), ("All files", "*.*")), initialdir=f"{path}/{input_path}")

    # if the user doesn't select a file, abort
    if paths_tuple is None or len(paths_tuple)==0:
        return
    
    input_video_paths.extend(paths_tuple)
    input_video_paths.reverse()

    # capture first frame and update video player gui
    if cap is not None and cap.isOpened():
        cap.release()
    cap = cv2.VideoCapture(input_video_paths[-1])
    video_selected = os.path.basename(input_video_paths[-1])
    video_label.config(text=f"Current Video Selected:\n\n {video_selected}", font=("Terminal", 20))
    success, preview_image = cap.read()
    if success:
        # Convert image from one color space to other
        opencv_preview_image = cv2.cvtColor(preview_image, cv2.COLOR_BGR2RGBA)

        # Capture the latest frame and transform to image
        captured_preview_image = Image.fromarray(opencv_preview_image)

        # Resize image
        captured_preview_image = captured_preview_image.resize((480, 270))

        # Convert captured image to photoimage
        photo_preview_image = ImageTk.PhotoImage(image=captured_preview_image)

        # Displaying photoimage in the label
        image_widget.photo_image = photo_preview_image

        # Configure image in the label
        image_widget.configure(image=photo_preview_image)
    if cap is not None and cap.isOpened():
        cap.release()
    # Reset frame number when opening new file
    frame_number = 0

    # Start processing the first video
    first_video = input_video_paths.pop()
    cap = cv2.VideoCapture(first_video)

    # create first sheet for first video
    sheet = global_workbook.active
    sheet.append(["Timesheet for:", f"{os.path.basename(first_video)}"])
    sheet.append(["Start Time (min:sec)", "End Time (min:sec)"])


def detect_bird(frame): #function for detecting the bird in a frame
    # do not try and write this variable to a file, it's not compatible
    processed = preprocess_frame(frame)
    if frame_number % frame_divisor == 0: # only check every frame divisible by preset number to save time
            # send processed frame to interpreter be checked for bird
            checked = check_frame(processed)
            # send checked frame to thresholding to see if confidence is high enough
            # if so, handle confidence stamp and timestamp
            thresholding(checked)

def preprocess_frame(frame):     # function for processing frames from capture
    if frame is None or frame.size == 0:
        return None
    
    resized_frame = cv2.resize(frame, (input_shape[1], input_shape[2]))
    normalized_frame = resized_frame / 255.0  
    return np.expand_dims(normalized_frame, axis=0)


def check_frame(processed):   # function for sending frame to interpreter to be checked for bird
        processed = np.float32(processed)
        interpreter.set_tensor(input_details[0]['index'], processed)
        interpreter.invoke()
        return interpreter.get_tensor(output_details[0]['index'])

def thresholding(checked_frame):    # function for thresholding based on confidence of model
    global confidence_percentage
    global timestamp_start
    global delay_started
    global delay_start_time
    global global_workbook

    # get confidence from model
    confidence = checked_frame[0][0]
    confidence_percentage = int(checked_frame[0][0] * 100)

    # update gui with new confidence percentage
    # if the gui percentage 
    if confidence >= confidence_threshold:
        confidence_percentage_label.config(text=f"{confidence_percentage}%", font=("Terminal", 20), foreground="green")
    else:
        confidence_percentage_label.config(text=f"{confidence_percentage}%", font=("Terminal", 20), foreground="black")

    # logic for timestamps
    if confidence > confidence_threshold: # check if confidence is above threshold, if it, start a timestamp if one hasn't been already  
        if timestamp_start is None:
            timestamp_start = cap.get(cv2.CAP_PROP_POS_MSEC) / 1000.0
        delay_started = False  # reset delay if confidence is back up
    else: # if we're below confidence_threshold
        if timestamp_start is not None:  # check if we had started a timestamp
            if not delay_started:  # start delay if it hasn't started
                delay_started = True
                delay_start_time = time.time()
            else:  # check if delay is over
                current_time = time.time()
                if current_time - delay_start_time >= delay_duration:
                    timestamp_end = cap.get(cv2.CAP_PROP_POS_MSEC) / 1000.0
                    minutes_start = int(timestamp_start // 60)
                    seconds_start = int(timestamp_start % 60)
                    minutes_end = int(timestamp_end // 60)
                    seconds_end = int(timestamp_end % 60)
                    timestamp_start_string = f"{minutes_start:02}:{seconds_start:02}"
                    timestamp_end_string = f"{minutes_end:02}:{seconds_end:02}"
                    sheet.append([timestamp_start_string, timestamp_end_string])
                    timestamp_start = None
                    delay_started = False

def set_model():    # function for setting model toggle when radio button is clicked
    global model_selected
    # get model selected from radio button value
    model = model_selection.get()
    # update path and label under video, write new selected model to file
    f = open('sprint-four/src/config.json', 'r+')
    settings = json.load(f)
    if model == 1:
        input_model_path = wren_model_path
        model_selected = "wren"
        settings["model_selected"] = "wren"
    else:
        input_model_path = warbler_model_path
        model_selected = "warbler"
        settings["model_selected"] = "warbler"
    model_label.config(text=f"Model Selected:\n\n {model_selected.capitalize()}")
    arrivals_departures_label.config(text=f"{model_selected.capitalize()} Arrivals & Departures", font=("Terminal", 20))
    f.seek(0)
    f.truncate()
    json.dump(settings, f)
    # Create a new interpreter with the selected model
    interpreter = tf.lite.Interpreter(model_path=input_model_path)
    interpreter.allocate_tensors()

def set_frame_skip_interval(): # update frame skip 

    def update_frame_skip(event):
        global frame_divisor
        frame_divisor = selected_frame_skip.get()
        f = open('sprint-four/src/config.json', 'r+')
        settings = json.load(f)
        settings["frame_divisor"] = str(frame_divisor)
        f.seek(0)
        f.truncate()
        json.dump(settings, f)
    selected_frame_skip = IntVar()
    selected_frame_skip.set(frame_divisor)

    frame_skip_window = Toplevel(root)
    frame_skip_window.geometry("300x100+200+0")
    frame_skip_window.title("Select Frame Divisor")
    frame_skip_label = tk.Label(frame_skip_window, text="Program will skip every frame not divisible by:")
    frame_skip_label.pack(side=TOP)
    frame_skip_select = ttk.Combobox(frame_skip_window, values=list(range(1,101)), state="readonly", textvariable=selected_frame_skip)
    frame_skip_select.pack(side=TOP)
    frame_skip_select.bind("<<ComboboxSelected>>", update_frame_skip)


def set_output_destination():
    # Write function later. Function should open up window to set output directory for video & spreadsheet
    pass

def toggle_playback():
    global playing
    global cap
    #if cap is None:
    #    print("No video selected. Select a video first.")
    #else:
    if playing:
        playing = False
        playback_button.config(image=play_image)
    else:   
        playing = True
        playback_button.config(image=pause_image)
        read_capture()

def show_timestamp_details(event):
    global video_selected

    index = arrivals_departures_text.index("@%s,%s" % (event.x, event.y))
    print(index)
    timestamp = arrivals_departures_text.get(index + " linestart", index + " lineend")
    start_time, end_time = map(lambda t: sum(int(x) * 60 ** i for i, x in enumerate(reversed(t.split(":")))), timestamp.split(" - "))

    # Get the total number of lines in the text widget
    total_lines = int(arrivals_departures_text.index('end-1c').split('.')[0])

    #print(check)

    video_names_dict = {}
    current_video = ""

    # Iterate through each line in the text widget
    for i in range(1, total_lines + 1):
        line_content = arrivals_departures_text.get(f"{i}.0", f"{i}.end").strip()  # Get the content of the line
        if "Timesheet for:" in line_content:
            sliced_line_content = line_content[17:]
            sliced_with_path = f"{input_path}{sliced_line_content}"
            video_names_dict[i] = {sliced_with_path}
    
    index_integer = int(float(index))
    #print(index_integer)

    for x in video_names_dict:
        if x < index_integer:
            video_set = video_names_dict[x]
            current_video = video_set.pop()
            #print(current_video)
            #print(type(current_video))


    #print(start_time)
    #print(end_time)
    
    #print(f"!!\n{video_selected}\n!!")
    #print(f"!!\n{input_video_path}\n!!")

    timestamp_cap = cv2.VideoCapture(current_video)
    timestamp_cap.set(cv2.CAP_PROP_POS_MSEC, start_time * 1000)
    
    new_window = Toplevel(root)
    new_window.title("Video with Timestamps")
    
    video_label = Label(new_window)
    video_label.pack()

    #timestamp_frame_number = 0 
    
    def update_frame():
        video_label
        ret, frame = timestamp_cap.read()
        #print(timestamp_cap.get(cv2.CAP_PROP_POS_MSEC))
        #print(end_time)
        if ret and timestamp_cap.get(cv2.CAP_PROP_POS_MSEC) <= end_time * 1000:
            
            resized_frame = cv2.resize(frame, (320, 240))
            
            cv2.putText(resized_frame, timestamp, (10, 30), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (255, 255, 255), 1)
            
            frame_rgb = cv2.cvtColor(resized_frame, cv2.COLOR_BGR2RGB)
            img = Image.fromarray(frame_rgb)
            imgtk = ImageTk.PhotoImage(image=img)
            video_label.imgtk = imgtk
            video_label.configure(image=imgtk)
            video_label.update()
            
            video_label.after(1, update_frame)
        else:
            timestamp_cap.release()
    def single_second_timestamp(): # special function for if timestamp is within a second, just show a frame from that second
        video_label
        ret, frame = timestamp_cap.read()
        resized_frame = cv2.resize(frame, (320, 240))
        
        cv2.putText(resized_frame, timestamp, (10, 30), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (255, 255, 255), 1)
        
        frame_rgb = cv2.cvtColor(resized_frame, cv2.COLOR_BGR2RGB)
        img = Image.fromarray(frame_rgb)
        imgtk = ImageTk.PhotoImage(image=img)
        video_label.imgtk = imgtk
        video_label.configure(image=imgtk)
        video_label.update()
        timestamp_cap.release()
    print(start_time)
    print(end_time)
    if start_time != end_time:
        update_frame()
    else:
        single_second_timestamp()

def read_capture():
    global playing
    global cap
    global input_video_path
    global frame_number

    if cap is not None and cap.isOpened():
        # Capture the video frame by frame
        _, frame = cap.read()

        # Increment frame number, update current frame in gui
        frame_number += 1
        current_frame_label.config(text=f"Current Frame: {frame_number}")

        # Send frame to detect_bird function to check for bird
        detect_bird(frame)

        # Convert image from one color space to other
        opencv_image = cv2.cvtColor(frame, cv2.COLOR_BGR2RGBA)

        # Capture the latest frame and transform to image
        captured_image = Image.fromarray(opencv_image)

        # Resize image
        captured_image = captured_image.resize((480, 270))

        # Convert captured image to photoimage
        photo_image = ImageTk.PhotoImage(image=captured_image)

        # Displaying photoimage in the label
        image_widget.photo_image = photo_image

        # Configure image in the label
        image_widget.configure(image=photo_image)

        # Update time position in gui
        time_raw = cap.get(cv2.CAP_PROP_POS_MSEC) / 1000.0
        minutes_raw = int(time_raw // 60)
        seconds_raw = int(time_raw % 60)
        time_position = f"{minutes_raw:02}:{seconds_raw:02}"
        time_position_label.config(text=f"Timestamp: {time_position}")

        # Update the timestamp label under "Arrivals & Departures"
        timestamps = "\n".join([f"{sheet.cell(row=i, column=1).value} - {sheet.cell(row=i, column=2).value}" for i in range(1, sheet.max_row+1)])
        arrivals_departures_text.delete(1.0, END)  # Clear the text widget
        arrivals_departures_text.insert(END, timestamps)
        #arrivals_departures_text.tag_add("clickable", "1.0", "end")

        # Assuming `formatting_lines` is a list containing the content of lines that should be excluded from having the "clickable" tag

        # Get the total number of lines in the text widget
        total_lines = int(arrivals_departures_text.index('end-1c').split('.')[0])

        #print(check)

        # Iterate through each line in the text widget
        for i in range(1, total_lines + 1):
            line_content = arrivals_departures_text.get(f"{i}.0", f"{i}.end").strip()  # Get the content of the line
            #print(line_content)
            if line_content not in formatting_lines:
                if not "Timesheet for:" in line_content:
                    # Apply the "clickable" tag to the line if it does not contain any of the formatting lines
                    arrivals_departures_text.tag_add("clickable", f"{i}.0", f"{i}.end")

        # Configure the "clickable" tag
        arrivals_departures_text.tag_configure("clickable", foreground="blue", underline=1)

        # Bind the "<Button-1>" event to the "show_timestamp_details" function for lines tagged as "clickable"
        arrivals_departures_text.tag_bind("clickable", "<Button-1>", show_timestamp_details)

        # Check if end of video is reached
        if cap.get(cv2.CAP_PROP_POS_FRAMES) < cap.get(cv2.CAP_PROP_FRAME_COUNT): # if not, keeping going
            if playing:
                image_widget.after(1, read_capture)
        else: # if it has been, load next video
            cap.release()  # Release the current video capture
            if len(input_video_paths) == 0: # if no more videos are left, finish
                print("All videos finished processing.")
                playback_button.invoke()
                image_widget.config(image=resized_ex_img)
                save_workbook()
            else: # if another video is left, update the video label gui element and open the capture to that new video
                video_selected = os.path.basename(input_video_paths[-1])
                sheet.append(["Timesheet for:", f"{os.path.basename(video_selected)}"])
                sheet.append(["Start Time (min:sec)", "End Time (min:sec)"])
                video_label.config(text=f"Current Video Selected:\n\n {video_selected}", font=("Terminal", 20))
                cap.open(input_video_paths.pop())
            if playing:
                image_widget.after(1, read_capture)

def save_workbook():
    global global_workbook
    
    try:
        # Extract the base name of the video file
        video_base_name = os.path.splitext(os.path.basename(input_video_path))[0]
        # Save the workbook with the video's base name as the file name
        #save_workbook_location = filedialog.asksaveasfilename(title="Save File As", filetypes=(("Excel File", "*.xlsx"), ("All files", "*.*")))
        save_workbook_location = filedialog.asksaveasfilename(title="Save Workbook As...", filetypes=(("Excel File", "*.xlsx"), ("All files", "*.*")), defaultextension=".xlsx", initialdir=f"{path}/{output_path}")
        #global_workbook.save(f"{output_path}/timestamps.xlsx")
        global_workbook.save(save_workbook_location)
        print("Workbook saved successfully")
    except Exception as e:
        print(f"Failed to save workbook: {e}")

if __name__ == "__main__":
    # Set up root window
    root = Tk()
    root.geometry("1200x600+0+0")
    root.title("Roc")

    # Create images to be used later.
    play_image = Image.open("config/assets/play.png")
    play_image = play_image.resize((25, 25))
    play_image = ImageTk.PhotoImage(play_image)
    pause_image = Image.open("config/assets/pause.png")
    pause_image = pause_image.resize((25, 25))
    pause_image = ImageTk.PhotoImage(pause_image)
    icon_image = Image.open("config/assets/icon.png")
    icon_image = icon_image.resize((25,25))
    icon_image = ImageTk.PhotoImage(icon_image)

    #set icon for root
    root.iconphoto(False, icon_image)

    # Set up left frame for video playback.
    left_frame = ttk.Frame(root, padding="3 3 12 12", width=500, height=800)
    left_frame.pack(side="left", anchor=NW, padx=25, pady=25)
    left_frame.pack_propagate(False)

    # Add a separator between the left and right frames
    separator = ttk.Separator(root, orient='vertical')
    separator.pack(side='left', fill='y', padx=5, pady=5)

    # Set up right frame for data display
    right_frame = ttk.Frame(root, padding="3 3 12 12", width=500, height=800)
    right_frame.pack(side="left", anchor=NW, padx=25, pady=25)

    # Set up Layout on left side
    # Placeholder image to represent video frame
    ex_img = Image.open("config/assets/video_example.png")
    ex_img = ex_img.resize((480, 270))
    resized_ex_img = ImageTk.PhotoImage(ex_img)
    image_widget = ttk.Label(left_frame, image=resized_ex_img)
    image_widget.pack(side=TOP, anchor=N)
    # playback button
    playback_button = ttk.Button(left_frame, image=play_image, command=toggle_playback)
    playback_button.pack(side=TOP, anchor=W, padx=200)
    # Create a label to display the currently selected model file
    model_label = ttk.Label(left_frame, text=f"Model Selected:\n\n {model_selected.capitalize()}", font=("Terminal", 20))
    model_label.pack(side=TOP, anchor=W, padx=10, pady=10)
    # Create a label to display the currently selected video file
    video_label = ttk.Label(left_frame, text=f"Current Video Selected:\n\n None", font=("Terminal", 20))
    video_label.pack(side=TOP, anchor=W, padx=10, pady=10)
    # test button
    # test_button = ttk.Button(left_frame, text="TEST", command=read_capture,) # logic not implemented
    # test_button.pack(side=TOP, anchor=W, padx=200)

    # Set up Layout on right side

    #Top Section
    # time position referred to as timestamp in gui for user ease
    time_position_label = ttk.Label(right_frame, text=f"Timestamp: {time_position}", font=("Terminal", 20))
    time_position_label.pack(side=TOP, anchor=NW)
    current_frame_label = ttk.Label(right_frame, text=f"Current Frame: {frame_number}", font=("Terminal", 20))
    current_frame_label.pack(side=TOP, anchor=NW)
    confidence_text_widget = Text(right_frame, height=20, width=100, border=False)
    confidence_text_widget.pack(side=TOP, anchor=NW)
    confidence_level_label = ttk.Label(confidence_text_widget, text=f"Confidence Level: ", font=("Terminal", 20))
    confidence_level_label.pack(side=LEFT, anchor=NW)
    confidence_percentage_label = ttk.Label(confidence_text_widget, text=f"{confidence_percentage}%", font=("Terminal", 20))
    confidence_percentage_label.pack(side=LEFT, anchor=NW)

    #Bottom Section (Split into separate frames?) Note that the top of this section of info is 200 below the top of the right frame
    arrivals_departures_label = ttk.Label(right_frame, text=f"{model_selected.capitalize()} Arrivals & Departures", font=("Terminal", 20), justify=LEFT)
    arrivals_departures_label.pack(side=TOP, anchor=NW, pady=20)

    # Scrollbar for arrival and departure
    arrivals_departures_scrollbar = Scrollbar(right_frame, orient=VERTICAL)
    arrivals_departures_scrollbar.pack(side=RIGHT, fill=Y)
    arrivals_departures_text = Text(right_frame, yscrollcommand=arrivals_departures_scrollbar.set, wrap=WORD, height=10)
    arrivals_departures_text.pack(side=TOP, anchor=NW, fill=BOTH, expand=True)
    arrivals_departures_scrollbar.config(command=arrivals_departures_text.yview)

    # Save the timestamps to the workbook
    save_button = ttk.Button(right_frame, text="Save Workbook", command=save_workbook)
    save_button.pack(side=TOP, anchor=NW, pady=10)

    # Create Menu
    menu = tk.Menu(root)

    # Create Open File Menu Command (code for full file menu commented out incase it needs to be re-implemented)
    #file_menu = tk.Menu(menu, tearoff=False)
    menu.add_command(label="Open File", command=open_files)
    #menu.add_cascade(label="File", menu=file_menu)

    # Create Model Menu
    model_menu = tk.Menu(menu, tearoff=False)
    model_selection = IntVar(value=model_int)
    model_menu.add_radiobutton(label="Wren",variable=model_selection, command=set_model, value=1)
    model_menu.add_radiobutton(label="Warbler",variable=model_selection, command=set_model, value=2)
    menu.add_cascade(label="Model", menu=model_menu)

    # Create Set Frame Skip Command
    menu.add_command(label="Set Frame Divisor", command=set_frame_skip_interval)

    # Add menu to root window
    root.config(menu=menu)

    root.mainloop()
