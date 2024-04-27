import os
os.environ['OPENCV_FFMPEG_READ_ATTEMPTS'] = "2147483647"
print(os.environ['OPENCV_FFMPEG_READ_ATTEMPTS'])
import cv2
import numpy as np
import tensorflow as tf
import openpyxl
import json
import time
from tkinter import Tk, filedialog, Button, Label, Entry, Text, Scrollbar
from tkinter import *
from tkinter import ttk
import tkinter as tk
from PIL import Image, ImageTk
from pathlib import Path
from collections import deque

path = Path(os.getcwd())
print(path)
#parent = path.parent.absolute()
#os.chdir(parent)
# define paths for project
# remember, normal current cwd is angry-birds-project
codebase_path = "sprint-four/codebase/"
models_path = "sprint-four/models/"
input_path = "sprint-four/input/"
output_path = "sprint-four/output/"


# load configuration settings
f = open(f'{codebase_path}/config.json', 'r+')
settings = json.load(f)
model_selected = settings["model_selected"]
wren_model_path = settings["wren_model_path"]
warbler_model_path = settings["warbler_model_path"]
input_video_path = settings["input_video_path"] # no longer needed, i believe?  (video files are selected in open_file)
output_video_path = settings["output_video_path"]
output_timestamps_path = settings["output_timestamps_path"]
frame_divisor = int(settings["frame_divisor"]) # Only frames with a frame number divisible by this number will be processed (1 for all frames, this is for optimization) 
confidence_threshold = float(settings["confidence_threshold"]) # confidence threshold should be between 0 and 1
model_int = 0

if model_selected == "wren":
    model_int = 1
    input_model_path = wren_model_path
else:
    model_int = 2
    input_model_path = warbler_model_path

# create interpreter and load with pre-trained model 
interpreter = tf.lite.Interpreter(model_path=input_model_path)
interpreter.allocate_tensors()


input_details = interpreter.get_input_details() # list of dictionaries, each dictionary has details about an input tensor
output_details = interpreter.get_output_details() # list of dictionaries, each dictionary has details about an input tensor
input_shape = input_details[0]['shape'] # array of shape of input tensor

# Declare global variables

# variable for whether or not the video is playback is paused. True for playing, False for Paused
playing = False
# variable for current frame being read. initialize when program is run and reset when a new video is loaded
frame_number = 0
cap = None
confidence_percentage = 0
timestamp_start = None
time_position = ""

# initialize variables for delay to avoid duplicate timestamps
# don't set to above zero, this code isn't working yet
delay_started = False
delay_start_time = None
delay_duration = 0  # delay duration in seconds (how long the bird needs to be missing to end a timestamp)

# Initialize variables for video selection
input_video_paths = deque()
current_video_index = 0
video_selected = ""

# open workbook for collecting timestamps to later output to excel file
# reserve first sheet for first video
global_workbook = openpyxl.Workbook()

def open_files(): #function to a set of files
    global input_video_paths
    global cap
    global frame_number
    global sheet

    # open system dialog to open video files
    paths_tuple = filedialog.askopenfilenames(title="Select Video Files", filetypes=(("MP4 files", "*.mp4"), ("All files", "*.*")), initialdir=f"{path}/{input_path}")

    print(f"!!\nPaths Tuple: {paths_tuple}\n!!")

    # if the user doesn't select a file, abort
    if paths_tuple is None or len(paths_tuple)==0:
        return
    
    input_video_paths.extend(paths_tuple)
    input_video_paths.reverse()

    # capture first frame and update video player gui
    if cap is not None and cap.isOpened():
        cap.release()
    cap = cv2.VideoCapture(input_video_paths[-1])
    video_selected = os.path.basename(input_video_paths[-1])
    video_label.config(text=f"Current Video Selected:\n\n {video_selected}", font=("Terminal", 20))
    success, preview_image = cap.read()
    if success:
        # Convert image from one color space to other
        opencv_preview_image = cv2.cvtColor(preview_image, cv2.COLOR_BGR2RGBA)

        # Capture the latest frame and transform to image
        captured_preview_image = Image.fromarray(opencv_preview_image)

        # Resize image
        captured_preview_image = captured_preview_image.resize((480, 270))

        # Convert captured image to photoimage
        photo_preview_image = ImageTk.PhotoImage(image=captured_preview_image)

        # Displaying photoimage in the label
        image_widget.photo_image = photo_preview_image

        # Configure image in the label
        image_widget.configure(image=photo_preview_image)
    if cap is not None and cap.isOpened():
        cap.release()
    # Reset frame number when opening new file
    frame_number = 0

    # Start processing the first video
    first_video = input_video_paths.pop()
    cap = cv2.VideoCapture(first_video)

    # create first sheet for first video
    sheet = global_workbook.active
    sheet.append(["Timesheet for:", f"{os.path.basename(first_video)}"])
    sheet.append(["Start Time (min:sec)", "End Time (min:sec)"])


def detect_bird(frame): #function for detecting the bird in a frame
    # do not try and write this variable to a file, it's not compatible
    processed = preprocess_frame(frame)
    if frame_number % frame_divisor == 0: # only check every frame divisible by preset number to save time
            # send processed frame to interpreter be checked for bird
            checked = check_frame(processed)
            # send checked frame to thresholding to see if confidence is high enough
            # if so, handle confidence stamp and timestamp
            thresholding(checked)

def preprocess_frame(frame):     # function for processing frames from capture
    if frame is None or frame.size == 0:
        return None
    
    resized_frame = cv2.resize(frame, (input_shape[1], input_shape[2]))
    normalized_frame = resized_frame / 255.0  
    return np.expand_dims(normalized_frame, axis=0)


def check_frame(processed):   # function for sending frame to interpreter to be checked for bird
        processed = np.float32(processed)
        interpreter.set_tensor(input_details[0]['index'], processed)
        interpreter.invoke()
        return interpreter.get_tensor(output_details[0]['index'])

def thresholding(checked_frame):    # function for thresholding based on confidence of model
    global confidence_percentage
    global timestamp_start
    global delay_started
    global delay_start_time
    global global_workbook

    # get confidence from model
    confidence = checked_frame[0][0]
    confidence_percentage = int(checked_frame[0][0] * 100)

    # update gui with new confidence percentage
    # if the gui percentage 
    if confidence >= confidence_threshold:
        confidence_percentage_label.config(text=f"{confidence_percentage}%", font=("Terminal", 20), foreground="green")
    else:
        confidence_percentage_label.config(text=f"{confidence_percentage}%", font=("Terminal", 20), foreground="black")

    # logic for timestamps
    if confidence > confidence_threshold: # check if confidence is above threshold, if it, start a timestamp if one hasn't been already  
        if timestamp_start is None:
            timestamp_start = cap.get(cv2.CAP_PROP_POS_MSEC) / 1000.0
        delay_started = False  # reset delay if confidence is back up
    else: # if we're below confidence_threshold
        if timestamp_start is not None:  # check if we had started a timestamp
            if not delay_started:  # start delay if it hasn't started
                delay_started = True
                delay_start_time = time.time()
            else:  # check if delay is over
                current_time = time.time()
                if current_time - delay_start_time >= delay_duration:
                    timestamp_end = cap.get(cv2.CAP_PROP_POS_MSEC) / 1000.0
                    minutes_start = int(timestamp_start // 60)
                    seconds_start = int(timestamp_start % 60)
                    minutes_end = int(timestamp_end // 60)
                    seconds_end = int(timestamp_end % 60)
                    timestamp_start_string = f"{minutes_start:02}:{seconds_start:02}"
                    timestamp_end_string = f"{minutes_end:02}:{seconds_end:02}"
                    sheet.append([timestamp_start_string, timestamp_end_string])
                    timestamp_start = None
                    delay_started = False
    #Testing delay
    #if delay_started:
        #print(f"Waiting for {delay_duration} seconds before recording timestamp...")

'''def process_next_video():
    global input_video_paths
    global cap
    global current_video_index

    if current_video_index < len(input_video_paths):
        input_video_path = input_video_paths[current_video_index]
        cap = cv2.VideoCapture(input_video_path)
        current_video_index += 1
    else:
        # All videos processed, do cleanup or display message
        print("All videos processed")
        # clear video queue
        input_video_paths = []
        # stop playing
        playback_button.invoke()
        return
    # Start processing the video
    read_capture()'''

def set_model():    # function for setting model toggle when radio button is clicked
    global model_selected
    # get model selected from radio button value
    model = model_selection.get()
    # update path and label under video, write new selected model to file
    f = open('sprint-four/codebase/config.json', 'r+')
    settings = json.load(f)
    if model == 1:
        input_model_path = wren_model_path
        model_selected = "wren"
        settings["model_selected"] = "wren"
    else:
        input_model_path = warbler_model_path
        model_selected = "warbler"
        settings["model_selected"] = "warbler"
    model_label.config(text=f"Model Selected:\n\n {model_selected.capitalize()}")
    arrivals_departures_label.config(text=f"{model_selected.capitalize()} Arrivals & Departures", font=("Terminal", 20))
    f.seek(0)
    f.truncate()
    json.dump(settings, f)
    # Create a new interpreter with the selected model
    interpreter = tf.lite.Interpreter(model_path=input_model_path)
    interpreter.allocate_tensors()

def set_frame_skip_interval():
    # Write function later. Function should open up window to set frame skip interval
    pass

def set_output_destination():
    # Write function later. Function should open up window to set output directory for video & spreadsheet
    pass

def toggle_playback():
    global playing
    global cap

    #if cap is None:
    #    print("No video selected. Select a video first.")
    #else:
    if playing:
        playing = False
        playback_button.config(image=play_image)
    else:   
        playing = True
        playback_button.config(image=pause_image)
        read_capture()

def read_capture():
    global playing
    global cap
    global input_video_path
    global frame_number

    if cap is not None and cap.isOpened():
        # Capture the video frame by frame
        _, frame = cap.read()

        # Increment frame number, update current frame in gui
        frame_number += 1
        current_frame_label.config(text=f"Current Frame: {frame_number}")

        # Send frame to detect_bird function to check for bird
        detect_bird(frame)

        # Convert image from one color space to other
        opencv_image = cv2.cvtColor(frame, cv2.COLOR_BGR2RGBA)

        # Capture the latest frame and transform to image
        captured_image = Image.fromarray(opencv_image)

        # Resize image
        captured_image = captured_image.resize((480, 270))

        # Convert captured image to photoimage
        photo_image = ImageTk.PhotoImage(image=captured_image)

        # Displaying photoimage in the label
        image_widget.photo_image = photo_image

        # Configure image in the label
        image_widget.configure(image=photo_image)

        # Update time position in gui
        time_raw = cap.get(cv2.CAP_PROP_POS_MSEC) / 1000.0
        minutes_raw = int(time_raw // 60)
        seconds_raw = int(time_raw % 60)
        time_position = f"{minutes_raw:02}:{seconds_raw:02}"
        time_position_label.config(text=f"Timestamp: {time_position}")

        # Update the timestamp label under "Arrivals & Departures"
        timestamps = "\n".join([f"{sheet.cell(row=i, column=1).value} - {sheet.cell(row=i, column=2).value}" for i in range(1, sheet.max_row+1)])
        arrivals_departures_text.delete(1.0, END)  # Clear the text widget
        arrivals_departures_text.insert(END, timestamps)

        # Check if end of video is reached
        if cap.get(cv2.CAP_PROP_POS_FRAMES) < cap.get(cv2.CAP_PROP_FRAME_COUNT): # if not, keeping going
            if playing:
                image_widget.after(1, read_capture)
        else: # if it has been, load next video
            cap.release()  # Release the current video capture
            if len(input_video_paths) == 0: # if no more videos are left, finish
                print("All videos finished processing.")
                playback_button.invoke()
                image_widget.config(image=resized_ex_img)
                save_workbook()
            else: # if another video is left, update the video label gui element and open the capture to that new video
                video_selected = os.path.basename(input_video_paths[-1])
                sheet.append(["Timesheet for:", f"{os.path.basename(video_selected)}"])
                sheet.append(["Start Time (min:sec)", "End Time (min:sec)"])
                video_label.config(text=f"Current Video Selected:\n\n {video_selected}", font=("Terminal", 20))
                cap.open(input_video_paths.pop())
            if playing:
                image_widget.after(1, read_capture)

def save_workbook():
    global global_workbook
    
    try:
        # Extract the base name of the video file
        video_base_name = os.path.splitext(os.path.basename(input_video_path))[0]
        # Save the workbook with the video's base name as the file name
        #save_workbook_location = filedialog.asksaveasfilename(title="Save File As", filetypes=(("Excel File", "*.xlsx"), ("All files", "*.*")))
        save_workbook_location = filedialog.asksaveasfilename(title="Save Workbook As...", filetypes=(("Excel File", "*.xlsx"), ("All files", "*.*")), defaultextension=".xlsx", initialdir=f"{path}/{output_path}")
        #global_workbook.save(f"{output_path}/timestamps.xlsx")
        global_workbook.save(save_workbook_location)
        print("Workbook saved successfully")
    except Exception as e:
        print(f"Failed to save workbook: {e}")

if __name__ == "__main__":
    # Set up root window
    root = Tk()
    root.geometry("1200x600+0+0")
    root.title("Roc")

    # Create images to be used later.
    play_image = Image.open("sprint-four/codebase/assets/play.png")
    play_image = play_image.resize((25, 25))
    play_image = ImageTk.PhotoImage(play_image)
    pause_image = Image.open("sprint-four/codebase/assets/pause.png")
    pause_image = pause_image.resize((25, 25))
    pause_image = ImageTk.PhotoImage(pause_image)

    # Set up left frame for video playback.
    left_frame = ttk.Frame(root, padding="3 3 12 12", width=500, height=800)
    left_frame.pack(side="left", anchor=NW, padx=25, pady=25)
    left_frame.pack_propagate(False)

    # Add a separator between the left and right frames
    separator = ttk.Separator(root, orient='vertical')
    separator.pack(side='left', fill='y', padx=5, pady=5)

    # Set up right frame for data display
    right_frame = ttk.Frame(root, padding="3 3 12 12", width=500, height=800)
    right_frame.pack(side="left", anchor=NW, padx=25, pady=25)

    # Set up Layout on left side
    # Placeholder image to represent video frame
    ex_img = Image.open("sprint-four/codebase/assets/video_example.png")
    ex_img = ex_img.resize((480, 270))
    resized_ex_img = ImageTk.PhotoImage(ex_img)
    image_widget = ttk.Label(left_frame, image=resized_ex_img)
    image_widget.pack(side=TOP, anchor=N)
    # playback button
    playback_button = ttk.Button(left_frame, image=play_image, command=toggle_playback)
    playback_button.pack(side=TOP, anchor=W, padx=200)
    # Create a label to display the currently selected model file
    model_label = ttk.Label(left_frame, text=f"Model Selected:\n\n {model_selected.capitalize()}", font=("Terminal", 20))
    model_label.pack(side=TOP, anchor=W, padx=10, pady=10)
    # Create a label to display the currently selected video file
    video_label = ttk.Label(left_frame, text=f"Current Video Selected:\n\n None", font=("Terminal", 20))
    video_label.pack(side=TOP, anchor=W, padx=10, pady=10)
    # test button
    # test_button = ttk.Button(left_frame, text="TEST", command=read_capture,) # logic not implemented
    # test_button.pack(side=TOP, anchor=W, padx=200)

    # Set up Layout on right side

    #Top Section
    # time position referred to as timestamp in gui for user ease
    time_position_label = ttk.Label(right_frame, text=f"Timestamp: {time_position}", font=("Terminal", 20))
    time_position_label.pack(side=TOP, anchor=NW)
    current_frame_label = ttk.Label(right_frame, text=f"Current Frame: {frame_number}", font=("Terminal", 20))
    current_frame_label.pack(side=TOP, anchor=NW)
    confidence_text_widget = Text(right_frame, height=20, width=100, border=False)
    confidence_text_widget.pack(side=TOP, anchor=NW)
    confidence_level_label = ttk.Label(confidence_text_widget, text=f"Confidence Level: ", font=("Terminal", 20))
    confidence_level_label.pack(side=LEFT, anchor=NW)
    confidence_percentage_label = ttk.Label(confidence_text_widget, text=f"{confidence_percentage}%", font=("Terminal", 20))
    confidence_percentage_label.pack(side=LEFT, anchor=NW)

    #Bottom Section (Split into separate frames?) Note that the top of this section of info is 200 below the top of the right frame
    arrivals_departures_label = ttk.Label(right_frame, text=f"{model_selected.capitalize()} Arrivals & Departures", font=("Terminal", 20), justify=LEFT)
    arrivals_departures_label.pack(side=TOP, anchor=NW, pady=20)

    # Scrollbar for arrival and departure
    arrivals_departures_scrollbar = Scrollbar(right_frame, orient=VERTICAL)
    arrivals_departures_scrollbar.pack(side=RIGHT, fill=Y)
    arrivals_departures_text = Text(right_frame, yscrollcommand=arrivals_departures_scrollbar.set, wrap=WORD, height=10)
    arrivals_departures_text.pack(side=TOP, anchor=NW, fill=BOTH, expand=True)
    arrivals_departures_scrollbar.config(command=arrivals_departures_text.yview)

    # Save the timestamps to the workbook
    save_button = ttk.Button(right_frame, text="Save Workbook", command=save_workbook)
    save_button.pack(side=TOP, anchor=NW, pady=10)

    # Create Menu
    menu = tk.Menu(root)

    # Create File Menu
    file_menu = tk.Menu(menu, tearoff=False)
    file_menu.add_command(label="Open File", command=open_files)
    recent_menu = tk.Menu(file_menu, tearoff=False)
    file_menu.add_cascade(label="Open Recent", menu=recent_menu) # (Logic for this command is not implemented yet.)
    recent_menu.add_command(label="Example Recent File") # placeholder for visual test
    menu.add_cascade(label="File", menu=file_menu)

    # Create Settings Menu
    settings_menu = tk.Menu(menu, tearoff=False)
    model_selection = IntVar(value=model_int)
    settings_menu.add_radiobutton(label="Wren",variable=model_selection, command=set_model,value=1)
    settings_menu.add_radiobutton(label="Warbler",variable=model_selection, command=set_model,value=2)
    #settings_menu.add_command(label="Select Model", command=set_model) # (Logic for this command is not implemented yet.)
    settings_menu.add_separator()
    settings_menu.add_command(label="Set Frame Skip Interval", command=set_frame_skip_interval) # (Logic for this command is not implemented yet.)
    settings_menu.add_command(label="Set Output Destination", command=set_output_destination)
    settings_menu.add_separator()
    settings_menu.add_checkbutton(label="Frame Skip") # Logic not implemented
    settings_menu.add_checkbutton(label="Output Video File") # Logic not implemented
    settings_menu.add_checkbutton(label="Output Timestamp Spreadsheet") # Logic not implemented
    menu.add_cascade(label="Settings", menu=settings_menu)
    # Add menu to root window
    root.config(menu=menu)

    root.mainloop()
